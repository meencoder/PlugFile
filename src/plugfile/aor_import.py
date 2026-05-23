"""Import the RRC GIS Viewer "Download Wells" export into AOR findings.

The RRC Public GIS Viewer has no API, but its **Download Wells** tool lets an
operator draw a radius (up to 2.5 mi) around a point and export every well
inside it — with operator/wellbore, completion, and plugging records — to a
CSV/XLSX file (GIS Viewer User Guide, p.20-22). That export is exactly the raw
data the area-of-review evaluation needs.

This module parses that export and turns it into the ``aor_findings`` list that
:func:`plugfile.aor.assess_aor` consumes:

  * tolerant header matching (the export's exact column names vary by RRC
    release, and ``.csv`` drops leading zeros so ``.xlsx`` is preferred);
  * haversine distance + compass bearing from the subject well, so the operator
    never measures by hand;
  * filters out wells already plugged (a filed W-3 means they're sealed —
    "ignore wells already plugged", guide p.20), dedupes multi-completion rows
    to the shallowest open zone (the primary conduit risk).

The resulting findings flow straight into ``assess_aor`` for isolation-plug
computation, so no AOR logic is duplicated here.

Usage::

    from plugfile.aor_import import parse_download_wells
    result = parse_download_wells(
        raw_bytes, filename="wells.xlsx",
        subject_lat=31.0184, subject_lon=-102.01, subject_td_ft=10500,
    )
    findings = result.findings           # -> aor_findings for assess_aor()
"""

from __future__ import annotations

import csv
import io
import math
import re
from dataclasses import dataclass, field
from typing import Any, Optional


# ---- geo helpers ------------------------------------------------------------

_EARTH_RADIUS_MI = 3958.7613


def haversine_mi(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance between two lat/long points, in miles."""
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * _EARTH_RADIUS_MI * math.asin(math.sqrt(a))


_COMPASS = ("N", "NE", "E", "SE", "S", "SW", "W", "NW")


def bearing(lat1: float, lon1: float, lat2: float, lon2: float) -> str:
    """Compass direction (N, NE, …) from point 1 toward point 2."""
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dl = math.radians(lon2 - lon1)
    y = math.sin(dl) * math.cos(p2)
    x = math.cos(p1) * math.sin(p2) - math.sin(p1) * math.cos(p2) * math.cos(dl)
    deg = (math.degrees(math.atan2(y, x)) + 360.0) % 360.0
    return _COMPASS[round(deg / 45.0) % 8]


# ---- header matching --------------------------------------------------------

def _norm(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (h or "").lower())


# Canonical field -> ordered alias fragments (matched against normalized headers).
_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "api":       ("apinumber", "apino", "api14", "api", "uwi"),
    "lat":       ("surfacelatitude", "latitude", "ddlat", "lat"),
    "lon":       ("surfacelongitude", "longitude", "ddlong", "lon", "long"),
    "well_type": ("welltype", "filingpurpose", "wellkind"),
    "depth":     ("totaldepth", "completiondepth", "plugbackdepth", "tvd", "depth", "td"),
    "plug_date": ("pluggingdate", "plugdate", "dateplugged", "pluggedate"),
    "status":    ("wellstatus", "status", "wellbore"),
    "lease":     ("leasename", "lease"),
    "well_no":   ("wellnumber", "wellno", "well"),
    "zone":      ("fieldname", "reservoir", "zone", "field"),
    "operator":  ("operatorname", "operator"),
}


def _resolve_columns(headers: list[str]) -> dict[str, int]:
    """Map each canonical field to a column index, tolerant of naming variants."""
    norm = [(_norm(h), i) for i, h in enumerate(headers)]
    out: dict[str, int] = {}
    for field_name, aliases in _FIELD_ALIASES.items():
        idx: Optional[int] = None
        for alias in aliases:                       # exact match first
            for h, i in norm:
                if h == alias:
                    idx = i
                    break
            if idx is not None:
                break
        if idx is None:                             # then substring
            for alias in aliases:
                for h, i in norm:
                    if alias in h:
                        idx = i
                        break
                if idx is not None:
                    break
        if idx is not None:
            out[field_name] = idx
    return out


# ---- output types -----------------------------------------------------------

@dataclass
class WellImportSummary:
    """Counts + diagnostics for one import."""
    total_rows: int = 0
    unique_wells: int = 0
    plugged_skipped: int = 0
    no_coordinates: int = 0
    out_of_radius: int = 0
    of_concern: int = 0
    radius_mi: float = 0.5
    distances_computed: bool = False


@dataclass
class WellImportResult:
    findings: list[dict[str, Any]] = field(default_factory=list)
    summary: WellImportSummary = field(default_factory=WellImportSummary)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        s = self.summary
        return {
            "findings": self.findings,
            "warnings": self.warnings,
            "summary": {
                "total_rows": s.total_rows,
                "unique_wells": s.unique_wells,
                "plugged_skipped": s.plugged_skipped,
                "no_coordinates": s.no_coordinates,
                "out_of_radius": s.out_of_radius,
                "of_concern": s.of_concern,
                "radius_mi": s.radius_mi,
                "distances_computed": s.distances_computed,
            },
        }


# ---- row reading ------------------------------------------------------------

def _rows_from_csv(content: bytes) -> list[list[str]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = content.decode("utf-8", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for r in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in r])
    wb.close()
    return rows


def _to_float(s: Any) -> Optional[float]:
    if s is None:
        return None
    try:
        v = float(str(s).strip().replace(",", ""))
        return v
    except (TypeError, ValueError):
        return None


# ---- main parser ------------------------------------------------------------

def parse_download_wells(
    content: bytes | str,
    *,
    filename: str = "",
    subject_lat: float | None = None,
    subject_lon: float | None = None,
    subject_td_ft: float | None = None,
    radius_mi: float = 0.5,
    subject_api: str | None = None,
) -> WellImportResult:
    """Parse a GIS Viewer 'Download Wells' export into AOR findings.

    Args:
        content:        Raw file bytes (or text for CSV).
        filename:       Used to detect format (.xlsx vs .csv).
        subject_lat/lon: Subject well coordinates; when given, each finding's
                        distance + direction are computed and wells beyond
                        ``radius_mi`` are dropped.
        subject_td_ft:  Subject well total depth (advisory; assess_aor uses it).
        radius_mi:      Area-of-review radius (default ½ mile).
        subject_api:    Subject well API, so the subject well itself is excluded.

    Returns:
        A :class:`WellImportResult` whose ``findings`` plug straight into
        :func:`plugfile.aor.assess_aor`.
    """
    if isinstance(content, str):
        content = content.encode("utf-8")

    result = WellImportResult()
    result.summary.radius_mi = radius_mi
    have_coords = subject_lat is not None and subject_lon is not None
    result.summary.distances_computed = have_coords

    # ---- read rows ----------------------------------------------------------
    try:
        if filename.lower().endswith(".xlsx") or content[:2] == b"PK":
            rows = _rows_from_xlsx(content)
        else:
            rows = _rows_from_csv(content)
    except Exception as exc:
        result.warnings.append(f"Could not read the export file: {exc}")
        return result

    if not rows:
        result.warnings.append("The export file is empty.")
        return result

    headers = rows[0]
    cols = _resolve_columns(headers)
    if "api" not in cols:
        result.warnings.append(
            "Could not find an API-number column in the export — is this the "
            "GIS Viewer 'Download Wells' file? Headers seen: "
            + ", ".join(h for h in headers[:12] if h)
        )
        return result

    data_rows = rows[1:]
    result.summary.total_rows = len(data_rows)

    subject_api_norm = re.sub(r"\D", "", subject_api or "")

    # ---- collapse rows -> one record per well (shallowest open zone) --------
    wells: dict[str, dict[str, Any]] = {}
    for row in data_rows:
        def cell(field_name: str) -> str:
            i = cols.get(field_name)
            return row[i].strip() if (i is not None and i < len(row) and row[i]) else ""

        api = cell("api")
        if not api:
            continue
        api_norm = re.sub(r"\D", "", api)
        if subject_api_norm and api_norm == subject_api_norm:
            continue  # don't flag the subject well against itself

        plug_date = cell("plug_date")
        status = cell("status").lower()
        plugged = bool(plug_date) or "plug" in status

        depth = _to_float(cell("depth"))
        lat = _to_float(cell("lat"))
        lon = _to_float(cell("lon"))

        rec = wells.get(api_norm)
        if rec is None:
            rec = {
                "api": api, "api_norm": api_norm, "plugged": plugged,
                "depth": depth, "lat": lat, "lon": lon,
                "lease": cell("lease"), "well_no": cell("well_no"),
                "zone": cell("zone"), "operator": cell("operator"),
                "well_type": cell("well_type"),
            }
            wells[api_norm] = rec
        else:
            rec["plugged"] = rec["plugged"] or plugged
            # keep the shallowest completion depth (primary conduit risk)
            if depth is not None and (rec["depth"] is None or depth < rec["depth"]):
                rec["depth"] = depth
            for k, v in (("lat", lat), ("lon", lon)):
                if rec[k] is None and v is not None:
                    rec[k] = v

    result.summary.unique_wells = len(wells)

    # ---- classify each well -------------------------------------------------
    for rec in wells.values():
        if rec["plugged"]:
            result.summary.plugged_skipped += 1
            continue

        distance_mi = direction = None
        if have_coords:
            if rec["lat"] is None or rec["lon"] is None:
                result.summary.no_coordinates += 1
                # keep it as a finding but flag the missing distance
            else:
                distance_mi = round(
                    haversine_mi(subject_lat, subject_lon, rec["lat"], rec["lon"]), 3
                )
                direction = bearing(subject_lat, subject_lon, rec["lat"], rec["lon"])
                if distance_mi > radius_mi:
                    result.summary.out_of_radius += 1
                    continue  # outside the AOR — drop it

        well_label = rec["api"]
        if rec["lease"]:
            wn = f" #{rec['well_no']}" if rec["well_no"] else ""
            well_label = f"{rec['api']} ({rec['lease']}{wn})"

        result.summary.of_concern += 1
        result.findings.append({
            "well_id": well_label,
            "zone_name": rec["zone"] or None,
            "depth_ft": rec["depth"],
            "distance_mi": distance_mi,
            "direction": direction,
            # requires_isolation is intentionally left unset — assess_aor()
            # infers it (penetrated zone within the subject TD).
        })

    if not have_coords:
        result.warnings.append(
            "Subject well coordinates were not available, so distances could "
            "not be computed — every imported well is treated as in-radius. "
            "Verify each against the GIS Viewer."
        )
    if result.summary.no_coordinates:
        result.warnings.append(
            f"{result.summary.no_coordinates} imported well(s) had no "
            "coordinates; distance/direction left blank for those."
        )
    if not result.findings:
        result.warnings.append(
            "No wells of concern found in the export (all plugged or outside "
            f"the {radius_mi}-mi radius)."
        )

    return result
