"""Generate the LinkedIn software-mining workbook for Validation Phase, step 2.

Creates `tools/linkedin_software_mining_template.xlsx` with four sheets:
  1. Instructions
  2. Search Queries (4 surfaces: Google site-search, LinkedIn-native, vendor
     company pages, LinkedIn Posts URLs)
  3. Mining Log (50 rows, dropdowns, conditional formatting, real Table)
  4. Summary (auto-tally formulas)

Why the queries look the way they do: LinkedIn's free-tier search has
disabled the AND/OR/NOT/parenthetical Boolean syntax that used to work,
so a query like `"Greasebook" AND "Texas"` now returns nothing. The
queries below are organized into four surfaces; Google site-search and
vendor company pages are the highest-yield options.

Run:
    python tools/build_linkedin_mining_xlsx.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


# (label, query/url, surface guidance)
SEARCH_QUERIES = [
    # --- Surface 1: Google site-search. Highest yield. Honors Booleans. ---
    ("Greasebook -- Google site-search",
     'site:linkedin.com/in/ "Greasebook" Texas',
     "Paste into Google. Click each profile result."),
    ("WolfePak -- Google site-search",
     'site:linkedin.com/in/ "WolfePak" (Texas OR Permian)',
     "Paste into Google. Click each profile result."),
    ("Quorum -- Google site-search",
     'site:linkedin.com/in/ "Quorum" "oil and gas" Texas',
     "Paste into Google. Click each profile result."),
    ("Enverus -- Google site-search",
     'site:linkedin.com/in/ "Enverus" Texas operator',
     "Paste into Google. Click each profile result."),
    ("PakEnergy -- Google site-search",
     'site:linkedin.com/in/ "PakEnergy" Texas',
     "Paste into Google. Click each profile result."),
    ("Plugging / P&A -- Google site-search",
     'site:linkedin.com/in/ "plugging" Texas operator',
     "Paste into Google. Click each profile result."),
    ("Posts mentioning Greasebook -- Google",
     'site:linkedin.com/posts "Greasebook" Texas',
     "Paste into Google. Surfaces public posts mentioning the tool."),
    # --- Surface 2: LinkedIn-native, simplified. Single keyword + UI filters.
    ("Greasebook -- LinkedIn search",
     "Greasebook",
     "Paste into LinkedIn search. UI: Location=Texas, Industry=Oil and Gas."),
    ("WolfePak -- LinkedIn search",
     "WolfePak",
     "Paste into LinkedIn search. UI: Location=Texas, Industry=Oil and Gas."),
    ("Quorum -- LinkedIn search",
     "Quorum oil and gas",
     "Paste into LinkedIn search. UI: Location=Texas."),
    ("Enverus -- LinkedIn search",
     "Enverus",
     "Paste into LinkedIn search. UI: Location=Texas, Industry=Oil and Gas."),
    ("PakEnergy -- LinkedIn search",
     "PakEnergy",
     "Paste into LinkedIn search. UI: Location=Texas, Industry=Oil and Gas."),
    # --- Surface 3: vendor company pages. Highest confidence per minute. ---
    ("Greasebook employees",
     "https://www.linkedin.com/company/greasebook/people/",
     "Open URL. Read Customer Success / Sales bios for operator names."),
    ("WolfePak employees",
     "https://www.linkedin.com/company/wolfepak-software/people/",
     "Open URL. Read Customer Success / Sales bios for operator names."),
    ("Quorum Software employees",
     "https://www.linkedin.com/company/quorum-software/people/",
     "Open URL. Read Customer Success / Sales bios for operator names."),
    ("Enverus employees",
     "https://www.linkedin.com/company/enverus/people/",
     "Open URL. Read Customer Success / Sales bios for operator names."),
    ("PakEnergy employees",
     "https://www.linkedin.com/company/pakenergy/people/",
     "Open URL. Read Customer Success / Sales bios for operator names."),
    # --- Surface 4: LinkedIn Posts search URLs (different result set). ---
    ("Greasebook -- LinkedIn Posts",
     "https://www.linkedin.com/search/results/content/?keywords=Greasebook",
     "Open URL. Apply Posted-by-1st-and-2nd-connections or Past-month filter."),
    ("Plugging Texas -- LinkedIn Posts",
     "https://www.linkedin.com/search/results/content/?keywords=plugging%20Texas%20operator",
     "Open URL. Catch operator-side complaints / RRC compliance posts."),
]

SOFTWARE_OPTIONS = [
    "Greasebook", "WolfePak", "Quorum", "Enverus", "PakEnergy",
    "Spreadsheet", "Other", "Unknown",
]

CONFIDENCE_OPTIONS = ["high", "med", "low"]

MINING_HEADERS = [
    "operator_name", "software_mentioned", "source_url",
    "role_title", "confidence", "notes", "date_added",
]

MINING_ROW_COUNT = 50


FONT_BASE = Font(name="Arial", size=11)
FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Arial", size=14, bold=True)
FONT_SUBTLE = Font(name="Arial", size=10, italic=True, color="666666")

FILL_HEADER = PatternFill("solid", start_color="1F2937")
FILL_BANDED = PatternFill("solid", start_color="F3F4F6")
FILL_HIGH = PatternFill("solid", start_color="D1FAE5")
FILL_MED = PatternFill("solid", start_color="FEF3C7")
FILL_LOW = PatternFill("solid", start_color="FEE2E2")

BORDER_THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def _apply_font_to_used_range(ws, font=FONT_BASE):
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            if cell.font is None or cell.font.name != "Arial":
                cell.font = font


def build_instructions(ws):
    ws.title = "Instructions"
    ws["A1"] = "LinkedIn Software-Mining Template"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = "Validation Phase -- Step 2 -- Plugfile"
    ws["A2"].font = FONT_SUBTLE

    body = [
        "",
        "GOAL",
        "Build a 50-row map of which Texas oil & gas operators use which "
        "operations software. Use this to (a) gauge the addressable wedge "
        "for Plugfile and (b) seed cold-outreach lists.",
        "",
        "BUDGET",
        "Spend ~3 hours total. If you hit 50 rows in 90 min, stop -- the "
        "next marginal row is worth less than the next marginal call.",
        "",
        "WORKFLOW",
        "1. Open the 'Search Queries' tab. There are four surfaces: Google "
        "site-search, LinkedIn-native (simplified), vendor company pages, "
        "and LinkedIn Posts URLs. Start with Google site-search and vendor "
        "company pages -- those have the highest yield-per-minute. "
        "LinkedIn-native search alone returns very little since LinkedIn "
        "disabled Boolean syntax for free accounts.",
        "2. For Google site-search: paste the query into Google, click "
        "each profile link in results. For vendor company pages: open the "
        "URL, click 'People', read each Customer Success / Sales bio for "
        "operator names they cover.",
        "3. For each hit, log a row in 'Mining Log' with operator, "
        "software, URL, role, confidence, and any notes worth remembering.",
        "4. The 'Summary' tab auto-tallies adoption -- check it as you go "
        "to spot which software is over- or under-represented.",
        "",
        "CONFIDENCE GUIDE",
        "high  -- direct quote from a current employee profile or post",
        "med   -- inferred from job title, tool listed in Skills, or "
        "company About",
        "low   -- third-party mention, conference photo, or vendor case study",
        "",
        "WHEN TO STOP",
        "* You hit 50 rows, OR",
        "* You spent 3 hours and have at least 30 rows, OR",
        "* You can name the top 3 most-adopted tools with >5 high-"
        "confidence rows each. Whichever comes first.",
        "",
        "WHAT TO DO WITH THE RESULT",
        "Hand the 50-row list to the fractional advisor (step 4) and to "
        "the first 2 expert-network calls (step 3). Ask: 'do these tools "
        "show up in your engagements?' This is how you turn a desk "
        "exercise into a validated wedge.",
    ]

    headers = {
        "GOAL", "BUDGET", "WORKFLOW", "CONFIDENCE GUIDE",
        "WHEN TO STOP", "WHAT TO DO WITH THE RESULT",
    }
    for i, line in enumerate(body, start=3):
        ws.cell(row=i, column=1, value=line)
        if line in headers:
            ws.cell(row=i, column=1).font = Font(
                name="Arial", size=11, bold=True
            )

    ws.column_dimensions["A"].width = 100
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = ALIGN_LEFT_TOP
    _apply_font_to_used_range(ws)


def build_search_queries(ws):
    ws.title = "Search Queries"
    ws["A1"] = "Search Queries"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = (
        "Four surfaces. Booleans no longer work on LinkedIn-free search; "
        "use Google site-search and vendor company pages instead."
    )
    ws["A2"].font = FONT_SUBTLE

    headers = ["#", "Label", "Query / URL", "How to use"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_LEFT_CENTER
        cell.border = BORDER_THIN

    for i, (label, query, surface) in enumerate(SEARCH_QUERIES, start=1):
        r = 4 + i
        ws.cell(row=r, column=1, value=i).alignment = ALIGN_CENTER
        ws.cell(row=r, column=2, value=label).alignment = ALIGN_LEFT_TOP
        ws.cell(row=r, column=3, value=query).alignment = ALIGN_LEFT_TOP
        ws.cell(row=r, column=4, value=surface).alignment = ALIGN_LEFT_TOP
        for c in (1, 2, 3, 4):
            ws.cell(row=r, column=c).border = BORDER_THIN
        if i % 2 == 0:
            for c in (1, 2, 3, 4):
                ws.cell(row=r, column=c).fill = FILL_BANDED

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 55

    for r in range(5, 5 + len(SEARCH_QUERIES)):
        ws.row_dimensions[r].height = 38

    _apply_font_to_used_range(ws)


def build_mining_log(ws):
    ws.title = "Mining Log"

    for col, h in enumerate(MINING_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_LEFT_CENTER
        cell.border = BORDER_THIN

    widths = {"A": 30, "B": 20, "C": 50, "D": 32, "E": 14, "F": 60, "G": 14}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    last_row = 1 + MINING_ROW_COUNT
    for r in range(2, last_row + 1):
        for c in range(1, len(MINING_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_LEFT_TOP

    sw_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(SOFTWARE_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,
        promptTitle="Software",
        prompt="Pick from the dropdown.",
    )
    sw_validation.add(f"B2:B{last_row}")
    ws.add_data_validation(sw_validation)

    conf_validation = DataValidation(
        type="list",
        formula1='"' + ",".join(CONFIDENCE_OPTIONS) + '"',
        allow_blank=True,
        showDropDown=False,
        promptTitle="Confidence",
        prompt="high = direct quote; med = inferred; low = third-party.",
    )
    conf_validation.add(f"E2:E{last_row}")
    ws.add_data_validation(conf_validation)

    date_validation = DataValidation(
        type="date",
        allow_blank=True,
        promptTitle="Date added",
        prompt="The day you logged this row.",
    )
    date_validation.add(f"G2:G{last_row}")
    ws.add_data_validation(date_validation)

    conf_range = f"E2:E{last_row}"
    ws.conditional_formatting.add(
        conf_range,
        CellIsRule(operator="equal", formula=['"high"'], fill=FILL_HIGH),
    )
    ws.conditional_formatting.add(
        conf_range,
        CellIsRule(operator="equal", formula=['"med"'], fill=FILL_MED),
    )
    ws.conditional_formatting.add(
        conf_range,
        CellIsRule(operator="equal", formula=['"low"'], fill=FILL_LOW),
    )

    ws.freeze_panes = "A2"

    table_ref = f"A1:{get_column_letter(len(MINING_HEADERS))}{last_row}"
    table = Table(displayName="MiningLog", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    _apply_font_to_used_range(ws)


def build_summary(ws):
    ws.title = "Summary"
    ws["A1"] = "Adoption Summary"
    ws["A1"].font = FONT_TITLE
    ws["A2"] = "Auto-tallies from the Mining Log. Refreshes when you save."
    ws["A2"].font = FONT_SUBTLE

    ws["A4"] = "Software"
    ws["B4"] = "Mentions"
    ws["C4"] = "% of logged"
    for col in (1, 2, 3):
        cell = ws.cell(row=4, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_LEFT_CENTER
        cell.border = BORDER_THIN

    log_range = f"'Mining Log'!B2:B{1 + MINING_ROW_COUNT}"
    total_logged_formula = f"COUNTA({log_range})"

    for i, sw in enumerate(SOFTWARE_OPTIONS, start=5):
        ws.cell(row=i, column=1, value=sw).alignment = ALIGN_LEFT_CENTER
        ws.cell(row=i, column=2,
                value=f'=COUNTIF({log_range},"{sw}")'
                ).alignment = ALIGN_CENTER
        ws.cell(
            row=i, column=3,
            value=f'=IF({total_logged_formula}=0,0,B{i}/{total_logged_formula})',
        ).number_format = "0.0%"
        for c in (1, 2, 3):
            ws.cell(row=i, column=c).border = BORDER_THIN
            if i % 2 == 0:
                ws.cell(row=i, column=c).fill = FILL_BANDED

    total_row = 5 + len(SOFTWARE_OPTIONS)
    ws.cell(row=total_row, column=1, value="Total logged").font = Font(
        name="Arial", size=11, bold=True
    )
    ws.cell(row=total_row, column=2,
            value=f"={total_logged_formula}").font = Font(
        name="Arial", size=11, bold=True
    )
    ws.cell(row=total_row, column=3, value=f"=B{total_row}/{MINING_ROW_COUNT}")
    ws.cell(row=total_row, column=3).number_format = "0.0%"
    for c in (1, 2, 3):
        ws.cell(row=total_row, column=c).border = BORDER_THIN

    conf_range = f"'Mining Log'!E2:E{1 + MINING_ROW_COUNT}"

    ws["E4"] = "Confidence"
    ws["F4"] = "Rows"
    ws["G4"] = "% of logged"
    for col in (5, 6, 7):
        cell = ws.cell(row=4, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_LEFT_CENTER
        cell.border = BORDER_THIN

    for i, conf in enumerate(CONFIDENCE_OPTIONS, start=5):
        ws.cell(row=i, column=5, value=conf).alignment = ALIGN_LEFT_CENTER
        ws.cell(row=i, column=6,
                value=f'=COUNTIF({conf_range},"{conf}")'
                ).alignment = ALIGN_CENTER
        ws.cell(
            row=i, column=7,
            value=f'=IF({total_logged_formula}=0,0,F{i}/{total_logged_formula})',
        ).number_format = "0.0%"
        for c in (5, 6, 7):
            ws.cell(row=i, column=c).border = BORDER_THIN

    progress_row = 5 + len(SOFTWARE_OPTIONS) + 3
    ws.cell(row=progress_row, column=1, value="Progress").font = Font(
        name="Arial", size=11, bold=True
    )
    ws.cell(row=progress_row, column=2, value=f"={total_logged_formula}")
    ws.cell(row=progress_row, column=3, value=f"of {MINING_ROW_COUNT}")
    ws.cell(row=progress_row + 1, column=1, value="Completion").font = Font(
        name="Arial", size=11, bold=True
    )
    ws.cell(row=progress_row + 1, column=2,
            value=f"={total_logged_formula}/{MINING_ROW_COUNT}"
            ).number_format = "0.0%"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    _apply_font_to_used_range(ws)


def build(output_path):
    wb = Workbook()
    build_instructions(wb.active)
    build_search_queries(wb.create_sheet("Search Queries"))
    build_mining_log(wb.create_sheet("Mining Log"))
    build_summary(wb.create_sheet("Summary"))
    wb.active = wb.sheetnames.index("Mining Log")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    here = Path(__file__).resolve().parent
    out = here / "linkedin_software_mining_template.xlsx"
    saved = build(out)
    print(f"Wrote {saved}")
    print(f"  sheets: Instructions, Search Queries, Mining Log, Summary")
    print(f"  rows ready: {MINING_ROW_COUNT}")
    print(f"  search-query rows: {len(SEARCH_QUERIES)}")


if __name__ == "__main__":
    main()
