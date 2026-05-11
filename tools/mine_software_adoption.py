"""Mine public web sources to populate software_mentioned for each operator row.

For every row in the Mining Log that has `operator_name` filled but
`software_mentioned` empty, this script calls Claude (claude-sonnet-4-6)
with the web_search tool and asks it to find PUBLIC mentions of which
oil & gas operations software the operator uses. Findings are written
back into the workbook with source URLs and conservative confidence ratings.

NOT A LINKEDIN SCRAPER. The web_search tool queries public search engines
(public results only, no auth, no LinkedIn API). LinkedIn pages may surface
when they appear in the public index, but no scraping or login happens.
All findings must be verified before you treat them as evidence -- treat
this as a research assistant, not a source of truth.

Confidence rules (enforced in the prompt + post-validated):
  high -- a named, current employee directly states use of the software
          (quote, profile, public post)
  med  -- vendor case study names the operator OR job posting requires the
          tool OR conference deck names operator + tool
  low  -- third-party mention, blog reference, or any inference

If no credible public mention is found, the row is marked
software=Unknown, confidence=low, notes="No public mentions found in
{N} searches" so the row is not silently skipped.

Cost: ~$0.05-0.15 per operator with web_search, depending on iterations.
Use --max-operators 3 to dry-run cheaply before processing the full sheet.

Rate-limit handling:
  - default --sleep 20s between operator calls (web_search inflates input
    tokens; on Tier 1 you have 30k input tokens per minute)
  - on a 429, retries up to --max-retries times, honoring the server's
    retry-after header when present, otherwise sleeping --rate-limit-pause

Usage:
    $env:ANTHROPIC_API_KEY = "sk-ant-..."   # PowerShell
    python tools/mine_software_adoption.py                    # process all empty
    python tools/mine_software_adoption.py --max-operators 3  # cap to 3
    python tools/mine_software_adoption.py --dry-run          # preview, no writes
    python tools/mine_software_adoption.py --operator "Acme"  # one operator only
    python tools/mine_software_adoption.py --confidence-floor med
        # only write rows that come back >= med (skip 'low' findings)
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = (
    Path(__file__).resolve().parent / "linkedin_software_mining_template.xlsx"
)

SOFTWARE_OPTIONS = {
    "Greasebook",
    "WolfePak",
    "Quorum",
    "Enverus",
    "PakEnergy",
    "Spreadsheet",
    "Other",
    "Unknown",
}
CONFIDENCE_OPTIONS = ["high", "med", "low"]
CONFIDENCE_RANK = {"high": 3, "med": 2, "low": 1}

SHEET_NAME = "Mining Log"

MODEL = "claude-sonnet-4-6"
MAX_TOKENS = 2048
MAX_WEB_SEARCHES_PER_OPERATOR = 5


SYSTEM_PROMPT = """You are a research assistant for an oil & gas compliance
software project. You research which operations software a given Texas oil
& gas operator uses, by searching public web sources only.

You MUST follow these rules:

1. Search for PUBLIC mentions only. Do not invent claims. If you cannot
   find evidence in 1-3 web searches, return software="Unknown".

2. Software MUST be one of: Greasebook, WolfePak, Quorum, Enverus,
   PakEnergy, Spreadsheet, Other, Unknown. If you find evidence of a
   software not in this list, use "Other" and explain in notes.

3. Confidence rules (apply STRICTLY):
   - "high" -- a named, currently-employed person at the operator directly
     states they use the software (LinkedIn quote, public post, named
     vendor testimonial).
   - "med" -- a vendor case study names the operator + product, OR a job
     posting at the operator requires the tool, OR a conference deck shows
     operator + product together.
   - "low" -- any third-party mention, blog reference, ranking list, or
     inference. Use "low" when in doubt.

4. Always return a source URL. If multiple sources, return the strongest
   one. If no source, return software="Unknown" with confidence="low".

5. Output STRICTLY this JSON shape inside <result>...</result> tags, with
   no other text after the closing tag:

<result>
{
  "software": "<one of the controlled values>",
  "url": "<source URL or empty string>",
  "role": "<job title of the source person, or empty string>",
  "confidence": "<high|med|low>",
  "notes": "<one sentence explaining what the source actually says>"
}
</result>
"""


def user_prompt_for(operator: str, hint: str) -> str:
    extra = f"\n\nContext from prior research notes: {hint}" if hint else ""
    return (
        f"Research which oil & gas operations software is used by the "
        f"operator: **{operator}** (Texas-based, upstream).\n\n"
        f"Search for public web mentions tying this operator to one of: "
        f"Greasebook, WolfePak, Quorum, Enverus, PakEnergy, or any other "
        f"named operations / accounting / regulatory software.\n\n"
        f"Return your finding in the <result>...</result> JSON shape "
        f"specified in the system prompt. Use 'Unknown' if 1-3 searches "
        f"surface no evidence.{extra}"
    )


@dataclass
class Finding:
    operator: str
    software: str
    url: str
    role: str
    confidence: str
    notes: str
    raw_response: str = ""
    error: str = ""

    def is_valid(self) -> bool:
        return (
            not self.error
            and self.software in SOFTWARE_OPTIONS
            and self.confidence in CONFIDENCE_OPTIONS
        )


@dataclass
class RunStats:
    processed: int = 0
    written: int = 0
    skipped_existing: int = 0
    skipped_low_confidence: int = 0
    errors: list = field(default_factory=list)


RESULT_RE = re.compile(r"<result>\s*(\{.*?\})\s*</result>", re.DOTALL)


def _extract_finding_json(text):
    m = RESULT_RE.search(text)
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except json.JSONDecodeError:
        return None


def _retry_after_seconds(exc):
    """Best-effort extract of retry-after from a RateLimitError response."""
    resp = getattr(exc, "response", None)
    if resp is None:
        return None
    headers = getattr(resp, "headers", None) or {}
    for key in ("retry-after", "Retry-After"):
        if key in headers:
            try:
                return float(headers[key])
            except (TypeError, ValueError):
                pass
    return None


def call_claude(
    client,
    operator,
    hint,
    max_retries=3,
    rate_limit_pause=65.0,
):
    """One round-trip with web_search enabled. Retries 429s with backoff."""
    attempt = 0

    while attempt <= max_retries:
        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=MAX_TOKENS,
                system=SYSTEM_PROMPT,
                tools=[
                    {
                        "type": "web_search_20250305",
                        "name": "web_search",
                        "max_uses": MAX_WEB_SEARCHES_PER_OPERATOR,
                    }
                ],
                messages=[
                    {"role": "user", "content": user_prompt_for(operator, hint)}
                ],
            )
            break
        except Exception as e:
            is_rate = type(e).__name__ == "RateLimitError" or "429" in str(e)
            if is_rate and attempt < max_retries:
                wait = _retry_after_seconds(e) or rate_limit_pause
                print(
                    f"   ...rate-limited; sleeping {wait:.0f}s "
                    f"(attempt {attempt + 1}/{max_retries})"
                )
                time.sleep(wait)
                attempt += 1
                continue
            return Finding(
                operator=operator,
                software="",
                url="",
                role="",
                confidence="",
                notes="",
                error=f"API error: {type(e).__name__}: {e}",
            )

    text_parts = [
        b.text for b in response.content if getattr(b, "type", "") == "text"
    ]
    raw = "\n".join(text_parts)
    parsed = _extract_finding_json(raw)
    if parsed is None:
        return Finding(
            operator=operator,
            software="",
            url="",
            role="",
            confidence="",
            notes="",
            raw_response=raw[:400],
            error="No <result> JSON block found in response",
        )

    return Finding(
        operator=operator,
        software=str(parsed.get("software", "")).strip(),
        url=str(parsed.get("url", "")).strip(),
        role=str(parsed.get("role", "")).strip(),
        confidence=str(parsed.get("confidence", "")).strip().lower(),
        notes=str(parsed.get("notes", "")).strip(),
        raw_response=raw[:400],
    )


def find_rows_to_process(ws, only_operator):
    """Return list of (row_num, operator_name, existing_notes) to mine."""
    out = []
    for r in range(2, ws.max_row + 1):
        op = ws.cell(row=r, column=1).value
        sw = ws.cell(row=r, column=2).value
        notes = ws.cell(row=r, column=6).value or ""
        if not op:
            continue
        if sw:
            continue
        if only_operator and only_operator.lower() not in str(op).lower():
            continue
        out.append((r, str(op), str(notes)))
    return out


def write_finding(ws, row, finding):
    ws.cell(row=row, column=2, value=finding.software)
    ws.cell(row=row, column=3, value=finding.url)
    ws.cell(row=row, column=4, value=finding.role)
    ws.cell(row=row, column=5, value=finding.confidence)
    existing = ws.cell(row=row, column=6).value or ""
    sep = " | " if existing and finding.notes else ""
    ws.cell(row=row, column=6, value=f"{existing}{sep}{finding.notes}")


def parse_args(argv):
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    p.add_argument(
        "--workbook", type=Path, default=DEFAULT_WORKBOOK,
        help=f"path to workbook (default: {DEFAULT_WORKBOOK.name})",
    )
    p.add_argument(
        "--max-operators", type=int, default=None,
        help="cap how many operators to process this run (cost guardrail)",
    )
    p.add_argument(
        "--operator", type=str, default=None,
        help="process only operators whose name contains this substring",
    )
    p.add_argument(
        "--confidence-floor",
        choices=CONFIDENCE_OPTIONS,
        default="low",
        help="only write rows >= this confidence (default: low = write all)",
    )
    p.add_argument(
        "--dry-run", action="store_true",
        help="show what would be written; do not modify the workbook",
    )
    p.add_argument(
        "--sleep", type=float, default=20.0,
        help=(
            "seconds to sleep between operator calls (default: 20.0 -- web "
            "search inflates input tokens, so a long pace prevents per-minute "
            "rate-limit errors on Tier 1 accounts)"
        ),
    )
    p.add_argument(
        "--max-retries", type=int, default=3,
        help="how many times to retry on a 429 (default: 3)",
    )
    p.add_argument(
        "--rate-limit-pause", type=float, default=65.0,
        help=(
            "seconds to sleep after a 429 if the server didn't suggest a "
            "retry-after (default: 65 -- just over the per-minute cap)"
        ),
    )
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(sys.argv[1:] if argv is None else argv)

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print(
            "ERROR: ANTHROPIC_API_KEY not set.\n"
            "  PowerShell: $env:ANTHROPIC_API_KEY = 'sk-ant-...'\n"
            "  cmd:        set ANTHROPIC_API_KEY=sk-ant-...",
            file=sys.stderr,
        )
        return 2

    try:
        import anthropic
    except ImportError:
        print(
            "ERROR: anthropic SDK not installed. Run:\n"
            "  pip install anthropic",
            file=sys.stderr,
        )
        return 2

    client = anthropic.Anthropic(api_key=api_key)

    if not args.workbook.exists():
        print(
            f"ERROR: {args.workbook} not found. Run:\n"
            f"  python tools/build_linkedin_mining_xlsx.py",
            file=sys.stderr,
        )
        return 2

    wb = load_workbook(args.workbook)
    ws = wb[SHEET_NAME]

    targets = find_rows_to_process(ws, args.operator)
    if args.max_operators:
        targets = targets[: args.max_operators]

    if not targets:
        print("No rows to process. Either everything is mined, or no rows seeded.")
        return 0

    print(
        f"Mining {len(targets)} operator(s) from {args.workbook.name} "
        f"(model={MODEL}, dry-run={args.dry_run}, "
        f"floor={args.confidence_floor}, sleep={args.sleep}s)"
    )
    print("-" * 72)

    stats = RunStats()
    floor_rank = CONFIDENCE_RANK[args.confidence_floor]

    for i, (row_num, op, hint) in enumerate(targets, start=1):
        print(f"[{i}/{len(targets)}] row {row_num}: {op}")
        finding = call_claude(
            client,
            op,
            hint,
            max_retries=args.max_retries,
            rate_limit_pause=args.rate_limit_pause,
        )
        stats.processed += 1

        if not finding.is_valid():
            err = finding.error or (
                f"invalid software/confidence: "
                f"{finding.software}/{finding.confidence}"
            )
            print(f"   SKIP -- {err}")
            if finding.raw_response:
                print(f"   raw: {finding.raw_response[:200]}")
            stats.errors.append(f"{op}: {err}")
            time.sleep(args.sleep)
            continue

        rank = CONFIDENCE_RANK[finding.confidence]
        if rank < floor_rank:
            print(
                f"   SKIP -- confidence={finding.confidence} below floor "
                f"({args.confidence_floor})"
            )
            stats.skipped_low_confidence += 1
            time.sleep(args.sleep)
            continue

        print(
            f"   {finding.software} | {finding.confidence} | "
            f"{finding.url[:60] or '(no url)'}"
        )
        print(f"   notes: {finding.notes[:120]}")

        if not args.dry_run:
            write_finding(ws, row_num, finding)
            wb.save(args.workbook)
            stats.written += 1

        time.sleep(args.sleep)

    print("-" * 72)
    print(
        f"Done. processed={stats.processed} "
        f"written={stats.written} "
        f"skipped_low_conf={stats.skipped_low_confidence} "
        f"errors={len(stats.errors)}"
    )
    if stats.errors:
        print("Errors:")
        for e in stats.errors:
            print(f"  - {e}")

    if args.dry_run:
        print("\n(dry-run -- workbook NOT modified)")

    return 0 if not stats.errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
