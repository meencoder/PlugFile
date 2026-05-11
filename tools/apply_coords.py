"""Coordinate calibration tool for the W-3 PDF overlay.

Two modes:

  EXPORT (default): write the current W3_COORDS and grid constants from
  pdf_export.py into an editable xlsx.

    python tools/apply_coords.py --export -o out/coords.xlsx

  APPLY: read an edited xlsx back, re-render calib.pdf so you can see the
  moved crosshairs, and optionally patch the source constants.

    # Just re-render calib.pdf from the edited sheet:
    python tools/apply_coords.py out/coords_edited.xlsx --calib out/calib2.pdf

    # Re-render AND patch pdf_export.py in-place:
    python tools/apply_coords.py out/coords_edited.xlsx --calib out/calib2.pdf --patch

Workflow
--------
1. python tools/apply_coords.py --export -o out/coords.xlsx
2. Open out/coords.xlsx in Google Sheets (File > Import), edit x/y values.
3. File > Download > xlsx.
4. python tools/apply_coords.py out/coords_edited.xlsx --calib out/calib2.pdf
5. Open calib2.pdf alongside the W-3 form; verify crosshairs sit in the right cells.
6. Repeat steps 2-5 until satisfied.
7. python tools/apply_coords.py out/coords_edited.xlsx --patch  (updates source)
8. Run pytest to confirm no regressions.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

# ---- resolve project paths -------------------------------------------------
_HERE = Path(__file__).resolve().parent
_REPO = _HERE.parent
sys.path.insert(0, str(_REPO / "src"))

from plugfile.pdf_export import (
    CASING_COL_X,
    CASING_ROW_Y,
    PERF_PAIR_X,
    PERF_ROW_Y,
    PLUG_COL_X,
    PLUG_ROW_Y,
    W3_COORDS,
    FieldCoord,
    render_calibration_overlay,
)

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# ---- style helpers ---------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
_EDITABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
_LOCKED_FILL  = PatternFill("solid", fgColor="F2F2F2")

_WHITE_BOLD = Font(bold=True, color="FFFFFF", size=10)
_BOLD = Font(bold=True, size=10)


def _h(ws, row, col, text):
    """Write a bold header cell."""
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _HEADER_FILL
    c.font = _WHITE_BOLD
    c.alignment = Alignment(horizontal="center")
    return c


def _sh(ws, row, col, text):
    """Write a section-label cell (lighter blue)."""
    c = ws.cell(row=row, column=col, value=text)
    c.fill = _SECTION_FILL
    c.font = _WHITE_BOLD
    return c


def _editable(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _EDITABLE_FILL
    c.alignment = Alignment(horizontal="right")
    return c


def _locked(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = _LOCKED_FILL
    return c


def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ---- EXPORT ----------------------------------------------------------------

def export_xlsx(dest: Path) -> None:
    wb = Workbook()

    _write_scalars_tab(wb.active)
    wb.active.title = "scalars"
    _write_plug_cols_tab(wb.create_sheet("plug_cols"))
    _write_plug_rows_tab(wb.create_sheet("plug_rows"))
    _write_casing_tab(wb.create_sheet("casing"))
    _write_perf_tab(wb.create_sheet("perf"))
    _write_instructions_tab(wb.create_sheet("HOW_TO_USE"))

    wb.save(dest)
    print(f"Exported {len(W3_COORDS)} scalar coords + grid constants to {dest}")


def _write_scalars_tab(ws) -> None:
    ws.freeze_panes = "A2"
    headers = ["field_name", "page\n(0=front\n1=back)", "x", "y",
               "max_width", "font_size", "description / notes"]
    for col, h in enumerate(headers, 1):
        _h(ws, 1, col, h)
        ws.row_dimensions[1].height = 36

    _FIELD_NOTES = {
        "api_number": "value drawn after the printed '42-' prefix",
        "rrc_district": "2-char district (01-10, 6E, 7B, 7C, 8A)",
        "field_name": "Section II - RRC field name",
        "lease_number": "RRC lease/ID number",
        "lease_name": "lease name",
        "well_number": "well number on lease",
        "county": "Texas county name",
        "operator_name": "legal operator name",
        "operator_address": "operator mailing address",
        "footage_ns": "Section II - feet from N/S line",
        "footage_ew": "Section II - feet from E/W line",
        "section_block_survey": "Section/Block/Survey",
        "latitude": "decimal degrees",
        "longitude": "decimal degrees",
        "total_depth_ft": "Section III - total measured depth",
        "spud_date": "YYYY-MM-DD",
        "completion_date": "YYYY-MM-DD",
        "plugging_date": "YYYY-MM-DD  (Section III + row 19)",
        "operator_signature_name": "Section X - cert rep name",
        "operator_title": "Section X - title",
        "certification_date": "YYYY-MM-DD",
        "cementing_company": "Section X - cementing co. name",
        "buqw_depth_ft": "page 2 Section VII - BUQW depth",
        "gau_letter_reference": "page 2 Section VII - GAU letter ref",
        "surface_restoration_narrative": "page 2 Section IX narrative text",
    }

    for row, (name, coord) in enumerate(W3_COORDS.items(), 2):
        _locked(ws, row, 1, name)
        _editable(ws, row, 2, coord.page)
        _editable(ws, row, 3, coord.x)
        _editable(ws, row, 4, coord.y)
        _editable(ws, row, 5, coord.max_width)
        _editable(ws, row, 6, coord.font_size)
        _locked(ws, row, 7, _FIELD_NOTES.get(name, ""))

    widths = [28, 12, 8, 8, 12, 12, 45]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)


def _write_plug_cols_tab(ws) -> None:
    _sh(ws, 1, 1, "PLUG_COL_X — center x of each of the 8 plug columns")
    ws.merge_cells("A1:C1")
    _h(ws, 2, 1, "col_index (1-8)")
    _h(ws, 2, 2, "center_x")
    _h(ws, 2, 3, "notes")
    for i, x in enumerate(PLUG_COL_X, 1):
        _locked(ws, i + 2, 1, i)
        _editable(ws, i + 2, 2, x)
        _locked(ws, i + 2, 3, f"PLUG #{i}")
    for col, w in [(1, 16), (2, 12), (3, 14)]:
        _set_col_width(ws, col, w)


def _write_plug_rows_tab(ws) -> None:
    _ROW_DESC = {
        "cement_date":      "*19  Cementing Date (uses plugging_date)",
        "hole_size_in":     "20   Size of Hole or Pipe (inches)",
        "drill_pipe_depth": "21   Depth to Bottom of Tubing/Drill Pipe",
        "sacks":            "*22  Sacks of Cement Used",
        "slurry_volume":    "*23  Slurry Volume Pumped (cu ft)",
        "calc_top":         "*24  Calculated Top of Plug (ft)",
        "measured_top":     "25   Measured Top of Plug (ft)",
        "slurry_weight":    "*26  Slurry Wt (# / Gal)  — cementer fills",
        "type_cement":      "*27  Type Cement — cementer fills",
    }
    _sh(ws, 1, 1, "PLUG_ROW_Y — y position of each plug-grid row")
    ws.merge_cells("A1:C1")
    _h(ws, 2, 1, "row_name")
    _h(ws, 2, 2, "y")
    _h(ws, 2, 3, "description")
    for i, (name, y) in enumerate(PLUG_ROW_Y.items(), 1):
        _locked(ws, i + 2, 1, name)
        _editable(ws, i + 2, 2, y)
        _locked(ws, i + 2, 3, _ROW_DESC.get(name, ""))
    for col, w in [(1, 20), (2, 10), (3, 50)]:
        _set_col_width(ws, col, w)


def _write_casing_tab(ws) -> None:
    _COL_DESC = {
        "size": "od_in — casing OD in inches",
        "weight": "weight_lb_per_ft",
        "put_in": "set_depth_ft — depth casing was set",
        "left_in": "depth left after plug-back (defaults to set_depth_ft)",
        "hole_size": "bit size of hole — not in schema yet; left blank",
    }
    _sh(ws, 1, 1, "CASING — row y positions (5 rows)")
    ws.merge_cells("A1:B1")
    _h(ws, 2, 1, "row_index (1-5)")
    _h(ws, 2, 2, "y")
    for i, y in enumerate(CASING_ROW_Y, 1):
        _locked(ws, i + 2, 1, i)
        _editable(ws, i + 2, 2, y)

    _sh(ws, 10, 1, "CASING — column x positions")
    ws.merge_cells("A10:C10")
    _h(ws, 11, 1, "col_name")
    _h(ws, 11, 2, "x")
    _h(ws, 11, 3, "description")
    for i, (name, x) in enumerate(CASING_COL_X.items(), 1):
        _locked(ws, i + 11, 1, name)
        _editable(ws, i + 11, 2, x)
        _locked(ws, i + 11, 3, _COL_DESC.get(name, ""))
    for col, w in [(1, 16), (2, 10), (3, 45)]:
        _set_col_width(ws, col, w)


def _write_perf_tab(ws) -> None:
    _sh(ws, 1, 1, "PERF — row y positions (5 rows; 2 from-to pairs per row = 10 perfs max)")
    ws.merge_cells("A1:B1")
    _h(ws, 2, 1, "row_index (1-5)")
    _h(ws, 2, 2, "y")
    for i, y in enumerate(PERF_ROW_Y, 1):
        _locked(ws, i + 2, 1, i)
        _editable(ws, i + 2, 2, y)

    _sh(ws, 10, 1, "PERF — from/to x positions (2 pairs per row)")
    ws.merge_cells("A10:D10")
    _h(ws, 11, 1, "pair_index (1-2)")
    _h(ws, 11, 2, "from_x")
    _h(ws, 11, 3, "to_x")
    _h(ws, 11, 4, "description")
    _PAIR_DESC = {1: "left pair (perfs 1,3,5,7,9)", 2: "right pair (perfs 2,4,6,8,10)"}
    for i, (from_x, to_x) in enumerate(PERF_PAIR_X, 1):
        _locked(ws, i + 11, 1, i)
        _editable(ws, i + 11, 2, from_x)
        _editable(ws, i + 11, 3, to_x)
        _locked(ws, i + 11, 4, _PAIR_DESC.get(i, ""))
    for col, w in [(1, 16), (2, 10), (3, 10), (4, 40)]:
        _set_col_width(ws, col, w)


def _write_instructions_tab(ws) -> None:
    ws.column_dimensions["A"].width = 80
    rows = [
        "W-3 PDF Overlay — Coordinate Calibration Sheet",
        "",
        "HOW TO USE",
        "----------",
        "1. Open out/calib.pdf alongside this sheet.",
        "   Each red crosshair on the PDF is labeled with a field name.",
        "   A crosshair sitting in the wrong cell means x/y needs adjusting.",
        "",
        "2. Edit the yellow cells on the other tabs to move crosshairs:",
        "   • scalars tab  — x/y for every scalar field (operator, dates, depths, etc.)",
        "   • plug_cols    — center_x for each of the 8 plug columns",
        "   • plug_rows    — y for each plug-grid row (*19 cementing date … *27 type cement)",
        "   • casing       — row y positions (5 rows) and column x positions",
        "   • perf         — row y positions and from/to x positions",
        "",
        "   PDF page coordinate system: origin at BOTTOM-LEFT, y increases UP.",
        "   Page is 612 wide × 792 tall (standard US Letter in PDF user-space units).",
        "   The template renders at 72 dpi so 1 pt ≈ 1/72 inch.",
        "",
        "3. File > Download > Microsoft Excel (.xlsx).",
        "",
        "4. Run:  python tools/apply_coords.py <your_edited_file.xlsx> --calib out/calib2.pdf",
        "   Open calib2.pdf to see the moved crosshairs.",
        "",
        "5. Repeat steps 2-4 until all crosshairs land in the right cells.",
        "",
        "6. Run:  python tools/apply_coords.py <your_edited_file.xlsx> --patch",
        "   This rewrites the constant blocks in src/plugfile/pdf_export.py.",
        "   Then run pytest to confirm all 238 tests still pass.",
        "",
        "TIPS",
        "----",
        "• On the W-3 form, y ≈ 780 is the very top; y ≈ 20 is the very bottom.",
        "• Typical row-to-row spacing is 12-13 pt. If a field is one row too high,",
        "  subtract ~12 from y.",
        "• Typical column spacing is 50-60 pt between plug columns.",
        "• Do not edit the gray (locked) cells — they are field names and descriptions.",
        "• The 'page' column: 0 = front (Sections I-X), 1 = back (Sections 31-42).",
    ]
    for i, text in enumerate(rows, 1):
        c = ws.cell(row=i, column=1, value=text)
        if i == 1:
            c.font = Font(bold=True, size=14)
        elif text in ("HOW TO USE", "TIPS"):
            c.font = Font(bold=True, size=11)


# ---- APPLY -----------------------------------------------------------------

def apply_xlsx(src: Path) -> tuple[dict, tuple, dict, tuple, dict, tuple, tuple]:
    """Read an edited xlsx and return updated constant dicts/tuples."""
    wb = load_workbook(src, data_only=True)

    coords = _read_scalars(wb["scalars"])
    plug_col_x = _read_plug_cols(wb["plug_cols"])
    plug_row_y = _read_plug_rows(wb["plug_rows"])
    casing_row_y, casing_col_x = _read_casing(wb["casing"])
    perf_row_y, perf_pair_x = _read_perf(wb["perf"])

    return coords, plug_col_x, plug_row_y, casing_row_y, casing_col_x, perf_row_y, perf_pair_x


def _read_scalars(ws) -> dict[str, FieldCoord]:
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, page, x, y, max_w, fs, *_ = row
        if name is None:
            continue
        out[str(name)] = FieldCoord(
            page=int(page), x=float(x), y=float(y),
            max_width=float(max_w), font_size=float(fs),
        )
    return out


def _read_plug_cols(ws) -> tuple[float, ...]:
    vals = []
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        idx, x = row
        if idx is None:
            continue
        vals.append((int(idx), float(x)))
    vals.sort(key=lambda t: t[0])
    return tuple(x for _, x in vals)


def _read_plug_rows(ws) -> dict[str, float]:
    out = {}
    for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
        name, y = row
        if name is None:
            continue
        out[str(name)] = float(y)
    return out


def _read_casing(ws) -> tuple[tuple[float, ...], dict[str, float]]:
    row_y = []
    for row in ws.iter_rows(min_row=3, max_row=7, max_col=2, values_only=True):
        idx, y = row
        if idx is None:
            continue
        row_y.append((int(idx), float(y)))
    row_y.sort(key=lambda t: t[0])

    col_x: dict[str, float] = {}
    for row in ws.iter_rows(min_row=12, max_col=2, values_only=True):
        name, x = row
        if name is None:
            continue
        col_x[str(name)] = float(x)

    return tuple(y for _, y in row_y), col_x


def _read_perf(ws) -> tuple[tuple[float, ...], tuple[tuple[float, float], ...]]:
    row_y = []
    for row in ws.iter_rows(min_row=3, max_row=7, max_col=2, values_only=True):
        idx, y = row
        if idx is None:
            continue
        row_y.append((int(idx), float(y)))
    row_y.sort(key=lambda t: t[0])

    pairs = []
    for row in ws.iter_rows(min_row=12, max_col=3, values_only=True):
        idx, from_x, to_x = row
        if idx is None:
            continue
        pairs.append((int(idx), float(from_x), float(to_x)))
    pairs.sort(key=lambda t: t[0])

    return tuple(y for _, y in row_y), tuple((fx, tx) for _, fx, tx in pairs)


# ---- PATCH SOURCE ----------------------------------------------------------

def patch_source(
    coords: dict,
    plug_col_x: tuple,
    plug_row_y: dict,
    casing_row_y: tuple,
    casing_col_x: dict,
    perf_row_y: tuple,
    perf_pair_x: tuple,
) -> None:
    """Rewrite the constant blocks in pdf_export.py in-place."""
    src_path = _REPO / "src" / "plugfile" / "pdf_export.py"
    source = src_path.read_text(encoding="utf-8")

    source = _replace_block(source, "W3_COORDS", _render_w3_coords(coords))
    source = _replace_block(source, "PLUG_COL_X", _render_plug_col_x(plug_col_x))
    source = _replace_block(source, "PLUG_ROW_Y", _render_plug_row_y(plug_row_y))
    source = _replace_block(source, "CASING_ROW_Y", _render_casing_row_y(casing_row_y))
    source = _replace_block(source, "CASING_COL_X", _render_casing_col_x(casing_col_x))
    source = _replace_block(source, "PERF_ROW_Y", _render_perf_row_y(perf_row_y))
    source = _replace_block(source, "PERF_PAIR_X", _render_perf_pair_x(perf_pair_x))

    src_path.write_text(source, encoding="utf-8")
    print(f"Patched {src_path}")


def _replace_block(source: str, var_name: str, new_assignment: str) -> str:
    """Replace the assignment block for `var_name` in `source`.

    Uses brace/paren depth tracking so the closing delimiter is always
    consumed, regardless of whether the block is one-line or multi-line.
    """
    lines = source.splitlines(keepends=True)

    # Find the line that starts the assignment
    start = None
    for i, line in enumerate(lines):
        if re.match(rf"^{re.escape(var_name)}\s*[:=]", line):
            start = i
            break
    if start is None:
        print(f"  WARNING: could not find '{var_name}' — skipped")
        return source

    # Walk forward tracking brace/paren depth until depth returns to 0
    depth = 0
    end = start
    for i in range(start, len(lines)):
        depth += lines[i].count("{") + lines[i].count("(") + lines[i].count("[")
        depth -= lines[i].count("}") + lines[i].count(")") + lines[i].count("]")
        if depth <= 0:
            end = i
            break

    before = "".join(lines[: start])
    after  = "".join(lines[end + 1 :])
    return before + new_assignment + "\n" + after


def _render_w3_coords(coords: dict) -> str:
    lines = ["W3_COORDS: dict[str, FieldCoord] = {"]
    for name, fc in coords.items():
        lines.append(
            f'    "{name}":{" " * max(1, 27 - len(name))}'
            f"FieldCoord({fc.page}, {fc.x}, {fc.y}, {fc.max_width}"
            + (f", {fc.font_size}" if fc.font_size != 9.0 else "")
            + "),"
        )
    lines.append("}")
    return "\n".join(lines)


def _render_plug_col_x(vals: tuple) -> str:
    inner = ", ".join(str(v) for v in vals)
    return f"PLUG_COL_X: tuple[float, ...] = ({inner})"


def _render_plug_row_y(d: dict) -> str:
    lines = ["PLUG_ROW_Y: dict[str, float] = {"]
    for k, v in d.items():
        lines.append(f'    "{k}":{" " * max(1, 22 - len(k))}{v},')
    lines.append("}")
    return "\n".join(lines)


def _render_casing_row_y(vals: tuple) -> str:
    inner = ", ".join(str(v) for v in vals)
    return f"CASING_ROW_Y: tuple[float, ...] = ({inner})"


def _render_casing_col_x(d: dict) -> str:
    lines = ["CASING_COL_X: dict[str, float] = {"]
    for k, v in d.items():
        lines.append(f'    "{k}":{" " * max(1, 12 - len(k))}{v},')
    lines.append("}")
    return "\n".join(lines)


def _render_perf_row_y(vals: tuple) -> str:
    inner = ", ".join(str(v) for v in vals)
    return f"PERF_ROW_Y: tuple[float, ...] = ({inner})"


def _render_perf_pair_x(vals: tuple) -> str:
    lines = ["PERF_PAIR_X: tuple[tuple[float, float], ...] = ("]
    for from_x, to_x in vals:
        lines.append(f"    ({from_x}, {to_x}),")
    lines.append(")")
    return "\n".join(lines)


# ---- CLI -------------------------------------------------------------------

def _cli():
    import argparse
    p = argparse.ArgumentParser(
        prog="apply_coords",
        description="Export / apply / patch W-3 overlay coordinates.",
    )
    p.add_argument("xlsx", nargs="?",
                   help="Edited coords xlsx (required for --calib and --patch)")
    p.add_argument("--export", action="store_true",
                   help="Export current constants to xlsx (default when no xlsx given)")
    p.add_argument("-o", "--output", default="out/coords.xlsx",
                   help="Destination for --export (default: out/coords.xlsx)")
    p.add_argument("--calib", metavar="PDF",
                   help="Re-render calibration overlay to this path")
    p.add_argument("--patch", action="store_true",
                   help="Patch src/plugfile/pdf_export.py constants in-place")
    a = p.parse_args()

    if a.export or a.xlsx is None:
        dest = Path(a.output)
        dest.parent.mkdir(parents=True, exist_ok=True)
        export_xlsx(dest)
        return

    src = Path(a.xlsx)
    if not src.exists():
        sys.exit(f"File not found: {src}")

    print(f"Reading {src} ...")
    (coords, plug_col_x, plug_row_y,
     casing_row_y, casing_col_x,
     perf_row_y, perf_pair_x) = apply_xlsx(src)

    if a.calib:
        import importlib
        import plugfile.pdf_export as _mod
        # Temporarily monkey-patch the module so render_calibration_overlay
        # uses the new coords from the xlsx.
        _mod.W3_COORDS = coords
        _mod.PLUG_COL_X = plug_col_x
        _mod.PLUG_ROW_Y = plug_row_y
        _mod.CASING_ROW_Y = casing_row_y
        _mod.CASING_COL_X = casing_col_x
        _mod.PERF_ROW_Y = perf_row_y
        _mod.PERF_PAIR_X = perf_pair_x

        calib_dest = Path(a.calib)
        calib_dest.parent.mkdir(parents=True, exist_ok=True)
        pdf = _mod.render_calibration_overlay()
        calib_dest.write_bytes(pdf)
        print(f"Calibration overlay -> {calib_dest} ({len(pdf):,} bytes)")

    if a.patch:
        patch_source(coords, plug_col_x, plug_row_y,
                     casing_row_y, casing_col_x, perf_row_y, perf_pair_x)
        print("Patched pdf_export.py — run pytest to verify.")


if __name__ == "__main__":
    _cli()
