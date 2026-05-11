"""Append rows to the LinkedIn software-mining workbook from the CLI.

Two modes:

  Single row (flag-driven):
    python tools/fill_mining_log.py \
      --operator "Acme Resources" \
      --software Greasebook \
      --url "https://www.linkedin.com/in/jane-doe-12345" \
      --role  "Production Superintendent" \
      --confidence high \
      --notes "quoted Greasebook in 2024 post about Permian field ops"

  Batch (JSON file -> many rows in one shot):
    python tools/fill_mining_log.py --batch tools/leads.json

  leads.json format -- a JSON list of objects with these keys:
    [
      {
        "operator": "Acme Resources",
        "software": "Greasebook",
        "url": "https://...",
        "role": "Production Superintendent",
        "confidence": "high",
        "notes": "..."
      },
      ...
    ]

The script appends to the first empty row in the Mining Log table, fills in
today's date if `date_added` is omitted, validates `software` and
`confidence` against the workbook's dropdowns, and saves in place.

Why this exists: to make logging a lead a single command, so the LinkedIn
hour stays in flow instead of context-switching to Excel each time.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = (
    Path(__file__).resolve().parent / "linkedin_software_mining_template.xlsx"
)

SOFTWARE_OPTIONS = {
    "Greasebook",
    "WolfePak",
    "Quorum",
    "Enverus",
    "PakEnergy",
    "Spreadsheet",
    "Other",
    "Unknown",
}
CONFIDENCE_OPTIONS = {"high", "med", "low"}
SHEET_NAME = "Mining Log"


def _validate(row: dict) -> None:
    if not row.get("operator"):
        raise ValueError("operator is required")
    sw = row.get("software", "")
    if sw and sw not in SOFTWARE_OPTIONS:
        raise ValueError(
            f"software '{sw}' must be one of: {sorted(SOFTWARE_OPTIONS)}"
        )
    conf = row.get("confidence", "")
    if conf and conf not in CONFIDENCE_OPTIONS:
        raise ValueError(
            f"confidence '{conf}' must be one of: {sorted(CONFIDENCE_OPTIONS)}"
        )


def _next_empty_row(ws) -> int:
    """First row with empty operator_name (column A) starting from row 2."""
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value in (None, ""):
            return r
    # all 50 rows full -- append a new one
    return ws.max_row + 1


def append_rows(workbook_path: Path, rows: list[dict]) -> list[int]:
    wb = load_workbook(workbook_path)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(
            f"sheet '{SHEET_NAME}' not found in {workbook_path} -- did you "
            f"run build_linkedin_mining_xlsx.py first?"
        )
    ws = wb[SHEET_NAME]

    written: list[int] = []
    today = dt.date.today().isoformat()

    for row in rows:
        _validate(row)
        r = _next_empty_row(ws)
        ws.cell(row=r, column=1, value=row["operator"])
        ws.cell(row=r, column=2, value=row.get("software", ""))
        ws.cell(row=r, column=3, value=row.get("url", ""))
        ws.cell(row=r, column=4, value=row.get("role", ""))
        ws.cell(row=r, column=5, value=row.get("confidence", ""))
        ws.cell(row=r, column=6, value=row.get("notes", ""))
        ws.cell(row=r, column=7, value=row.get("date_added") or today)
        written.append(r)

    wb.save(workbook_path)
    return written


def parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Append rows to the LinkedIn software-mining workbook."
    )
    p.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"path to the workbook (default: {DEFAULT_WORKBOOK})",
    )
    p.add_argument("--operator", help="operator / company name")
    p.add_argument(
        "--software",
        choices=sorted(SOFTWARE_OPTIONS),
        help="software mentioned",
    )
    p.add_argument("--url", help="LinkedIn source URL")
    p.add_argument("--role", help="role / title of the person")
    p.add_argument(
        "--confidence",
        choices=sorted(CONFIDENCE_OPTIONS),
        help="confidence in the mention",
    )
    p.add_argument("--notes", default="", help="freeform notes")
    p.add_argument(
        "--date",
        dest="date_added",
        help="ISO date (defaults to today)",
    )
    p.add_argument(
        "--batch",
        type=Path,
        help="JSON file with a list of row dicts (overrides single-row flags)",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)

    if args.batch is not None:
        rows = json.loads(args.batch.read_text(encoding="utf-8"))
        if not isinstance(rows, list):
            raise SystemExit(f"{args.batch} must contain a JSON list")
    else:
        if not args.operator:
            raise SystemExit(
                "--operator is required (or pass --batch <file.json>)"
            )
        rows = [
            {
                "operator": args.operator,
                "software": args.software or "",
                "url": args.url or "",
                "role": args.role or "",
                "confidence": args.confidence or "",
                "notes": args.notes,
                "date_added": args.date_added,
            }
        ]

    written = append_rows(args.workbook, rows)
    print(f"Appended {len(written)} row(s) to {args.workbook}")
    for r in written:
        print(f"  row {r}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
